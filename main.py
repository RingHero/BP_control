from numpy import *
import numpy as np
import matplotlib.pyplot as plt
import struct
import socket

import rtde_control

import random
import xlrd
import xlwt
from xlutils.copy import copy
from random import randint
from openpyxl import load_workbook
import rtde_receive

#左上角点1  右上角2  左下角3  右下角4  上边5  下边6  左边7  右边8  正碰撞9
def Transfer_Matrix(theta,d,a,alpha):
    #用于获取坐标系i ii到坐标系i − 1 i-1i−1的通用变换矩阵
    T = np.array([[cos(theta),-sin(theta)*cos(alpha),sin(theta)*sin(alpha),a*cos(theta)],
                  [sin(theta), cos(theta)*cos(alpha) , -cos(theta)*sin(alpha), a*sin(theta)],
                  [0, sin(alpha), cos(alpha), d],
                  [0, 0, 0, 1]])
    return T
def Forward_Kinematics(theta,d,a,alpha):
    #正运动学直接建模
    T01 = Transfer_Matrix(theta[0], d[0], a[0], alpha[0])
    T12 = Transfer_Matrix(theta[1], d[1], a[1], alpha[1])
    T23 = Transfer_Matrix(theta[2], d[2], a[2], alpha[2])
    T34 = Transfer_Matrix(theta[3], d[3], a[3], alpha[3])
    T45 = Transfer_Matrix(theta[4], d[4], a[4], alpha[4])
    T56 = Transfer_Matrix(theta[5], d[5], a[5], alpha[5])
    T = T01.dot(T12).dot(T23).dot(T34).dot(T45).dot(T56)
    #print(T00)
    return T

def sigmoid(z):
    h = 1. / (1 + np.exp(-z))
    return h
def de_sigmoid(h):
    return h * (1 - h)
# 无激活函数（全1矩阵）前反向传递函数
def no_active(z):
    h = z
    return h
def de_no_active(h):
    return np.ones(h.shape)
# 推演函数
def fead_forward(datas, layers):
    input_layers = []
    for i in range(len(layers)):
        layer = layers[i]
        if i == 0:
            inputs = datas
            z = np.dot(inputs, layer["w"]) + layer["b"]
            h = layer["act_fun"](z)
            input_layers.append(inputs)
        else:
            inputs = h
            z = np.dot(inputs, layer["w"]) + layer["b"]
            h = layer["act_fun"](z)
            input_layers.append(inputs)
    output = np.argmax(h)+1
    return output
def predict(F,model):
    trait = np.array([F[3] / F[2], F[4] / F[2], F[5] / F[2]])
    res = fead_forward(trait, model)
    return res
def write_exl(matrix):
    wb = xlwt.Workbook(encoding='ascii')
    ws = wb.add_sheet("first")

    for i in range(100):
        ws.write(i,0,random.randint(75,85))

    wb.save("D:\Working park\ew.xls")



class NET:
    def __init__(self,modelpath):
        self.model_path = modelpath
        self.model = np.load(self.model_path,allow_pickle= True)
    def print_model(self):
        print("模型如下")
        print(self.model)
class CSqQueue:

    def __init__(self):
        self.__MaxSize = 10
        self.data = [None] * self.__MaxSize
        self.front = self.rear = 0

    def empty(self):
        if self.front == self.rear:
            return True
        return False

    def enough(self):
        if (self.rear + 1) % self.__MaxSize == self.front:
            return True
        return False

    def push(self, e):
        assert (self.rear + 1) % self.__MaxSize != self.front
        self.rear = (self.rear + 1) % self.__MaxSize
        self.data[self.rear] = e

    def pop(self):
        assert not self.empty()
        self.front = (self.front + 1) % self.__MaxSize
        return self.data[self.front]

    def gethead(self):
        assert not self.empty()
        return self.data[(self.front + 1) % self.__MaxSize]

def data_recv(recv):
    data = recv.recv(1000)
    fx = struct.unpack('f', data[6:10])[0]
    fy = struct.unpack('f', data[10:14])[0]
    fz = struct.unpack('f', data[14:18])[0]
    mx = struct.unpack('f', data[18:22])[0]
    my = struct.unpack('f', data[22:26])[0]
    mz = struct.unpack('f', data[26:30])[0]
    F = [fx, fy, fz, mx, my, mz]
    return F

#传感器IP及端口
IP_ADDR = '192.168.0.109'
PORT1 = 30005

#ur5IP及端口
ur5_HOST = '192.168.0.100'
PORT2 = 30003

#加载模型
model = np.load(r"D:\Working park\model2.npy",allow_pickle=True)

#创建连接socket接口
s = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
b = socket.socket(socket.AF_INET,socket.SOCK_STREAM)

#连接
s.connect((IP_ADDR,PORT1))

#查看链接地址
sendData = "AT+EIP=?\r\n"
s.send(sendData.encode())
recvData = bytearray(s.recv(1000))
print(recvData)

#查看解耦矩阵
sendData = "AT+DCPMA6X8=?\r\n"
s.send(sendData.encode())



recvData = bytearray(s.recv(1000))
print(recvData)

#计算单位
sendData = "AT+DCPCU=?\r\n"
s.send(sendData.encode())
recvData = bytearray(s.recv(1000))
print(recvData)

#设置采样频率
set_update_rate = "AT+SMPF=100\r\n"
s.send(set_update_rate.encode())
recvData = bytearray(s.recv(1000))
print(recvData)

#矩阵运算单位
set_compute_unit = "AT+DCPCU=MVPV\r\n"
s.send(set_compute_unit.encode())

#上传数据格式
set_recieve_format = "AT+SGDM=(A01,A02,A03,A04,A05,A06);E;1;(WMA:1)\r\n"
s.send(set_recieve_format.encode())
recvData = bytearray(s.recv(1000))
print(recvData)

#连续上传数据包
get_data_stream = "AT+GSD\r\n"
s.send(get_data_stream.encode())

print("开始接收\r\n")
times = 0

#excel_init
wb = xlwt.Workbook(encoding='ascii')
ws = wb.add_sheet("first")
ws.write(0, 0, "Fx")
ws.write(0, 1, "Fy")
ws.write(0, 2, "Fz")
ws.write(0, 3, "Mx")
ws.write(0, 4, "My")
ws.write(0, 5, "Mz")
ws.write(0,6,"Type")

#kalman_para
hat_value = [0,0,0,0,0,0]
hat_var = [0.01,0.01,0.01,0.01,0.01,0.01]
read_var = [0.02,0.02,0.02,0.02,0.02,0.02]
kalman = [0,0,0,0,0,0]

#queue_init
queue_z = CSqQueue()

once = 0
total = 0
totalF = np.array([0,0,0,0,0,0])
F0 = [2.7179,-1.2820,-3.7647]
T0 = [0.0780,-0.0524,-0.2077]
P0 = [0.0383,-0.0434,0.0169]
rtde_r = rtde_receive.RTDEReceiveInterface("192.168.0.100")
while False:
    data = s.recv(1000)
    fx = struct.unpack('f',data[6:10])[0]
    fy = struct.unpack('f',data[10:14])[0]
    fz = struct.unpack('f',data[14:18])[0]
    mx = struct.unpack('f',data[18:22])[0]
    my = struct.unpack('f',data[22:26])[0]
    mz = struct.unpack('f',data[26:30])[0]

    #theta = -45.906708425463981616441938365535#°
    #rtde_r = rtde_receive.RTDEReceiveInterface("192.168.0.100")
    joint_q = rtde_r.getActualQ()
    #FK = Forward_Kinematics(joint_q, d=[89.159, 0, 0, 109.15, 94.65, 82.30], a=[0, -425.00, -392.25, 0, 0, 0],alpha=[np.pi / 2, 0, 0, np.pi / 2, -np.pi / 2, 0])
    #R31 = FK[2][0]
    #R32 = FK[2][1]
    #R33 = FK[2][2]


    if once ==0:

        F = np.array([fx, fy, fz, mx, my, mz])
        totalF = totalF + F
        once +=1



    if times==5:
        total += 1
        times = 0
        F = np.array([fx, fy, fz, mx, my, mz])

        joint_q = rtde_r.getActualQ()
        FK = Forward_Kinematics(joint_q, d=[89.159, 0, 0, 109.15, 94.65, 82.30], a=[0, -425.00, -392.25, 0, 0, 0],alpha=[np.pi / 2, 0, 0, np.pi / 2, -np.pi / 2, 0])
        R31 = FK[2][0]
        R32 = FK[2][1]
        R33 = FK[2][2]
        G = (-7.6787) * np.array(
            [-R31 * cos(-0.801) - R32 * sin(-0.801),
             -R32 * cos(-0.801) + R31 * sin(-0.801), -R33])
        Tg = np.array([G[2]*P0[1]-G[1]*P0[2],G[0]*P0[2]-G[2]*P0[0],G[1]*P0[0]-G[0]*P0[1]])
        Fxyz = np.array([fx, fy, fz])
        Txyz = np.array([mx,my,mz])
        totalF = totalF + F
        print(np.around(F, 5))
        F_real = Fxyz-F0-G
        T_real = Txyz-T0-Tg
        print("接触力：",np.round(F_real,5))
        print("力矩:",np.round(T_real,5))
        if F[2]<-3:
            #res = predict(F,model)
            #print("结果为：", res)
            pass
        else:
            print("结果为：未碰撞")

        '''
        ws.write(total, 0, F[0])
        ws.write(total, 1, F[1])
        ws.write(total, 2, F[2])
        ws.write(total, 3, F[3])
        ws.write(total, 4, F[4])
        ws.write(total, 5, F[5])
        ws.write(total, 6, 9)
        
        ws.write(total, 6, hat_value[0])
        ws.write(total, 7, hat_value[1])
        ws.write(total, 8, hat_value[2])
        ws.write(total, 9, hat_value[3])
        ws.write(total, 10, hat_value[4])
        ws.write(total, 11, hat_value[5])
        '''

    times +=1


    if total == 500:
        #print(totalF/500)
        #wb.save("D:\iDAS R&D20190730\正碰撞\hit4.xls")
        #break
        pass

exl = xlrd.open_workbook(r"C:\Users\11374\Desktop\测试.xls")
table = exl.sheets()[0]
nrows = table.nrows
n=100
#记录六维力以及电机角度，为了解出重力补偿
while False:
    exl = xlrd.open_workbook(r"C:\Users\11374\Desktop\测试.xls")
    table = exl.sheets()[0]
    nrows = table.nrows
    wbook = copy(exl)
    wsheet = wbook.get_sheet(0)
    print(nrows)
    Fs = data_recv(s)
    joints = rtde_r.getActualQ()
    datas = Fs+joints
    print(datas)
    for i in range(len(datas)):
        wsheet.write(nrows,i,datas[i])

    wbook.save(r"C:\Users\11374\Desktop\测试.xls")  #
    n-=1
for i in range(10):
    data_recv(s)
while n>0:
    for i in range(1):
        data_recv(s)
    Fs = data_recv(s)
    joints = rtde_r.getActualQ()
    datas = Fs+joints
    exl = xlrd.open_workbook(r"C:\Users\11374\Desktop\力.xls")
    table = exl.sheets()[8]
    nrows = table.nrows
    wbook = copy(exl)
    wsheet = wbook.get_sheet(8)
    print(nrows)
    print(datas)
    for j in range(len(datas)):
        wsheet.write(nrows, j, datas[j])
    wbook.save(r"C:\Users\11374\Desktop\力.xls")
    n-=1
